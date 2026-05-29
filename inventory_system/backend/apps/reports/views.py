from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework.response import Response
from django.utils import timezone
from django.http import HttpResponse
from datetime import date, timedelta
from decimal import Decimal
import io


def ok(data=None, msg='', code=status.HTTP_200_OK):
    return Response({'success': True, 'data': data or {}, 'message': msg, 'errors': {}}, status=code)

def err(errors=None, msg='', code=status.HTTP_400_BAD_REQUEST):
    return Response({'success': False, 'data': {}, 'message': msg, 'errors': errors or {}}, status=code)


def check_report_permission(user):
    return user.is_admin or user.is_manager


def build_daily_report(report_date):
    from apps.products.models import Variant
    from apps.stock.models import StockEntry, DailyLedger
    from apps.stock.stock_engine import get_live_stock
    from django.db.models import Sum

    variants = list(Variant.objects.filter(is_active=True).select_related('product'))
    v_ids = [v.id for v in variants]

    from apps.stock.stock_engine import bulk_get_live_stock
    live_stock_map = bulk_get_live_stock(v_ids, None, report_date)

    ledgers = {l.variant_id: l for l in DailyLedger.objects.filter(variant_id__in=v_ids, date=report_date)}

    in_entries = {e['variant_id']: e['t'] for e in StockEntry.objects.filter(
        variant_id__in=v_ids, timestamp__date=report_date, entry_type='IN', is_approved=True
    ).values('variant_id').annotate(t=Sum('quantity'))}

    out_entries = {e['variant_id']: e['t'] for e in StockEntry.objects.filter(
        variant_id__in=v_ids, timestamp__date=report_date, entry_type='OUT', is_approved=True
    ).values('variant_id').annotate(t=Sum('quantity'))}

    rows = []
    for v in variants:
        total_in = in_entries.get(v.id, Decimal('0'))
        total_out = out_entries.get(v.id, Decimal('0'))
        ledger = ledgers.get(v.id)
        opening = ledger.opening_stock if ledger else Decimal('0')
        closing = ledger.closing_stock
        live = live_stock_map.get(v.id, Decimal('0'))
        rows.append({
            'product': v.product.name,
            'variant': f'{v.size} {v.flavour} ({v.sku})'.strip(),
            'opening': float(opening),
            'total_in': float(total_in),
            'total_out': float(total_out),
            'live_stock': float(live),
            'closing': float(closing) if closing is not None else None,
        })
    return rows


class DailyReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not check_report_permission(request.user):
            return err(msg='Permission denied.', code=status.HTTP_403_FORBIDDEN)
        date_str = request.query_params.get('date', str(timezone.now().date()))
        try:
            report_date = date.fromisoformat(date_str)
        except ValueError:
            return err(msg='Invalid date format.')
        rows = build_daily_report(report_date)
        return ok(data={'date': date_str, 'rows': rows})


class WeeklyReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not check_report_permission(request.user):
            return err(msg='Permission denied.', code=status.HTTP_403_FORBIDDEN)
        from apps.products.models import Variant
        from apps.stock.models import StockEntry
        from django.db.models import Sum

        week_str = request.query_params.get('week')
        today = timezone.now().date()
        if week_str:
            try:
                year, week = week_str.split('-W')
                from_date = date.fromisocalendar(int(year), int(week), 1)
            except Exception:
                return err(msg='Invalid week format. Use YYYY-Www')
        else:
            from_date = today - timedelta(days=today.weekday())
        to_date = from_date + timedelta(days=6)

        rows = []
        variants = list(Variant.objects.filter(is_active=True).select_related('product'))
        v_ids = [v.id for v in variants]

        in_entries = {e['variant_id']: e['t'] for e in StockEntry.objects.filter(
            variant_id__in=v_ids, timestamp__date__range=[from_date, to_date], entry_type='IN', is_approved=True
        ).values('variant_id').annotate(t=Sum('quantity'))}

        out_entries = {e['variant_id']: e['t'] for e in StockEntry.objects.filter(
            variant_id__in=v_ids, timestamp__date__range=[from_date, to_date], entry_type='OUT', is_approved=True
        ).values('variant_id').annotate(t=Sum('quantity'))}

        for v in variants:
            total_in = in_entries.get(v.id, Decimal('0'))
            total_out = out_entries.get(v.id, Decimal('0'))
            if float(total_in) == 0 and float(total_out) == 0:
                continue
            rows.append({
                'product': v.product.name,
                'variant': f'{v.size} {v.flavour} ({v.sku})'.strip(),
                'total_in': float(total_in),
                'total_out': float(total_out),
                'net': float(total_in) - float(total_out),
            })
        return ok(data={'from_date': str(from_date), 'to_date': str(to_date), 'rows': rows})


class MonthlyReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not check_report_permission(request.user):
            return err(msg='Permission denied.', code=status.HTTP_403_FORBIDDEN)
        from apps.products.models import Variant
        from apps.stock.models import StockEntry
        from django.db.models import Sum
        import calendar

        month_str = request.query_params.get('month', str(timezone.now().date())[:7])
        try:
            year, month = map(int, month_str.split('-'))
            from_date = date(year, month, 1)
            _, last_day = calendar.monthrange(year, month)
            to_date = date(year, month, last_day)
        except Exception:
            return err(msg='Invalid month format. Use YYYY-MM')

        rows = []
        variants = list(Variant.objects.filter(is_active=True).select_related('product'))
        v_ids = [v.id for v in variants]

        in_entries = {e['variant_id']: e['t'] for e in StockEntry.objects.filter(
            variant_id__in=v_ids, timestamp__date__range=[from_date, to_date], entry_type='IN', is_approved=True
        ).values('variant_id').annotate(t=Sum('quantity'))}

        out_entries = {e['variant_id']: e['t'] for e in StockEntry.objects.filter(
            variant_id__in=v_ids, timestamp__date__range=[from_date, to_date], entry_type='OUT', is_approved=True
        ).values('variant_id').annotate(t=Sum('quantity'))}

        for v in variants:
            total_in = in_entries.get(v.id, Decimal('0'))
            total_out = out_entries.get(v.id, Decimal('0'))
            if float(total_in) == 0 and float(total_out) == 0:
                continue
            rows.append({
                'product': v.product.name,
                'variant': f'{v.size} {v.flavour} ({v.sku})'.strip(),
                'total_in': float(total_in),
                'total_out': float(total_out),
                'net': float(total_in) - float(total_out),
            })
        return ok(data={'month': month_str, 'from_date': str(from_date), 'to_date': str(to_date), 'rows': rows})


class MovementReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not check_report_permission(request.user):
            return err(msg='Permission denied.', code=status.HTTP_403_FORBIDDEN)
        from apps.stock.models import StockEntry
        from apps.stock.serializers import StockEntrySerializer

        variant_id = request.query_params.get('variant_id')
        from_date = request.query_params.get('from')
        to_date = request.query_params.get('to')
        location = request.query_params.get('location')

        qs = StockEntry.objects.select_related('variant__product', 'location', 'logged_by').filter(is_approved=True)
        if variant_id:
            qs = qs.filter(variant_id=variant_id)
        if from_date:
            qs = qs.filter(timestamp__date__gte=from_date)
        if to_date:
            qs = qs.filter(timestamp__date__lte=to_date)
        if location:
            qs = qs.filter(location_id=location)
        qs = qs.order_by('-timestamp')
        return ok(data={'results': StockEntrySerializer(qs, many=True).data, 'count': qs.count()})


class LowStockReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not check_report_permission(request.user):
            return err(msg='Permission denied.', code=status.HTTP_403_FORBIDDEN)
        from apps.products.models import Variant
        from apps.stock.stock_engine import get_live_stock
        today = timezone.now().date()
        result = []
        variants = list(Variant.objects.filter(is_active=True).select_related('product'))
        v_ids = [v.id for v in variants]
        from apps.stock.stock_engine import bulk_get_live_stock
        try:
            live_stock_map = bulk_get_live_stock(v_ids, None, today)
        except Exception as e:
            import logging
            logging.getLogger('django').error(f"Error bulk stock prefetch in LowStockReportView: {e}")
            live_stock_map = {}

        for v in variants:
            live = float(live_stock_map.get(v.id, 0))
            if live <= v.reorder_point:
                result.append({
                    'variant_id': v.id,
                    'sku': v.sku,
                    'product': v.product.name,
                    'size': v.size,
                    'flavour': v.flavour,
                    'live_stock': live,
                    'reorder_point': v.reorder_point,
                    'reorder_qty': v.reorder_qty,
                    'deficit': v.reorder_point - live,
                })
        return ok(data={'results': result, 'count': len(result)})


class CustomerBalanceView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not check_report_permission(request.user):
            return err(msg='Permission denied.', code=status.HTTP_403_FORBIDDEN)
        from apps.stock.models import Order
        
        customer_name = request.query_params.get('customer_name')
        qs = Order.objects.exclude(status=Order.STATUS_CANCELLED)
        if customer_name:
            qs = qs.filter(customer_name__icontains=customer_name)

        orders = qs.prefetch_related('items').all()
        balances = {}
        for order in orders:
            c = order.customer_name
            if c not in balances:
                balances[c] = {
                    'customer_name': c,
                    'ordered_cases': 0, 'delivered_cases': 0, 'pending_cases': 0,
                    'ordered_qty': 0, 'delivered_qty': 0, 'pending_qty': 0
                }
            
            for item in order.items.all():
                balances[c]['ordered_cases'] += (item.cases or 0)
                balances[c]['ordered_qty'] += float(item.quantity or 0)
                if order.status == Order.STATUS_DELIVERED:
                    balances[c]['delivered_cases'] += (item.cases or 0)
                    balances[c]['delivered_qty'] += float(item.quantity or 0)
                else:
                    balances[c]['pending_cases'] += (item.cases or 0)
                    balances[c]['pending_qty'] += float(item.quantity or 0)

        rows = list(balances.values())
        return ok(data={'results': rows})


class ExpiringReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not check_report_permission(request.user):
            return err(msg='Permission denied.', code=status.HTTP_403_FORBIDDEN)
        from apps.stock.models import StockEntry
        days = int(request.query_params.get('days', 30))
        cutoff = timezone.now().date() + timedelta(days=days)
        qs = StockEntry.objects.filter(
            expiry_date__isnull=False,
            expiry_date__lte=cutoff,
            expiry_date__gte=timezone.now().date(),
        ).select_related('variant__product', 'location').order_by('expiry_date')
        result = [{
            'variant': f'{e.variant.product.name} {e.variant.size} {e.variant.flavour}'.strip(),
            'sku': e.variant.sku,
            'batch_number': e.batch_number,
            'expiry_date': str(e.expiry_date),
            'quantity': float(e.quantity),
            'location': e.location.name,
            'days_until_expiry': (e.expiry_date - timezone.now().date()).days,
        } for e in qs]
        return ok(data={'results': result, 'count': len(result)})


class ForecastReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, variant_id):
        if not check_report_permission(request.user):
            return err(msg='Permission denied.', code=status.HTTP_403_FORBIDDEN)
        from apps.products.models import Variant
        from apps.stock.models import StockEntry
        from apps.stock.stock_engine import get_live_stock
        from django.db.models import Sum

        try:
            v = Variant.objects.get(pk=variant_id)
        except Variant.DoesNotExist:
            return err(msg='Variant not found.', code=status.HTTP_404_NOT_FOUND)

        today = timezone.now().date()
        from_date = today - timedelta(days=30)

        # Bulk fetch all daily OUT entries for the last 30 days in a single database query
        outs = (
            StockEntry.objects.filter(
                variant=v,
                entry_type='OUT',
                timestamp__date__range=[from_date, today],
                is_approved=True
            )
            .values('timestamp__date')
            .annotate(t=Sum('quantity'))
        )
        out_map = {o['timestamp__date']: o['t'] for o in outs}

        daily_out = []
        for i in range(30):
            d = from_date + timedelta(days=i)
            daily_out.append(float(out_map.get(d, Decimal('0'))))

        avg_daily = sum(daily_out) / 30 if any(daily_out) else 0
        live = float(get_live_stock(v.id, None, today))
        days_remaining = int(live / avg_daily) if avg_daily > 0 else None

        return ok(data={
            'variant_id': v.id,
            'sku': v.sku,
            'avg_daily_consumption': round(avg_daily, 3),
            'live_stock': live,
            'days_remaining': days_remaining,
            'daily_out_last_30': daily_out,
        })


class ExportReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, report_type):
        if not check_report_permission(request.user):
            return err(msg='Permission denied.', code=status.HTTP_403_FORBIDDEN)

        fmt = request.query_params.get('format', 'pdf')
        if report_type == 'daily':
            date_str = request.query_params.get('date', str(timezone.now().date()))
            try:
                report_date = date.fromisoformat(date_str)
            except ValueError:
                return err(msg='Invalid date format.')
            rows = build_daily_report(report_date)
            title = f'Daily Stock Report - {date_str}'
            headers = ['Product', 'Variant', 'Opening', 'IN', 'OUT', 'Live Stock', 'Closing']
            data_rows = [[r['product'], r['variant'], r['opening'], r['total_in'],
                          r['total_out'], r['live_stock'], r['closing'] or '-'] for r in rows]
        else:
            return err(msg='Report type not supported for export yet.')

        if fmt == 'excel':
            return generate_excel(title, headers, data_rows)
        else:
            return generate_pdf(title, headers, data_rows)


def generate_excel(title, headers, rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append([title])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
    return response


def generate_pdf(title, headers, rows):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles['Title']),
        Spacer(1, 12),
    ]
    table_data = [headers] + [[str(c) for c in r] for r in rows]
    t = Table(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{title}.pdf"'
    return response


class ForecastView(APIView):
    """
    GET /api/reports/forecast/<variant_id>/
    Computes 30-day moving average of daily OUT qty, estimates days remaining
    and projected stockout/reorder date.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, variant_id):
        from apps.products.models import Variant
        from apps.stock.models import StockEntry
        from apps.stock.stock_engine import get_live_stock
        from django.db.models import Sum

        try:
            variant = Variant.objects.select_related('product').get(pk=variant_id)
        except Variant.DoesNotExist:
            return err(msg='Variant not found.', code=status.HTTP_404_NOT_FOUND)

        today = timezone.now().date()
        thirty_days_ago = today - timedelta(days=30)

        # Get daily OUT quantities for the last 30 days
        daily_outs = (
            StockEntry.objects
            .filter(
                variant_id=variant_id,
                entry_type='OUT',
                is_approved=True,
                timestamp__date__gte=thirty_days_ago,
                timestamp__date__lte=today,
            )
            .values('timestamp__date')
            .annotate(daily_out=Sum('quantity'))
            .order_by('timestamp__date')
        )

        total_out = sum(d['daily_out'] for d in daily_outs)
        active_days = max(len(daily_outs), 1)
        avg_daily_consumption = float(total_out) / active_days

        current_stock = float(get_live_stock(variant_id, None, today))
        days_remaining = int(current_stock / avg_daily_consumption) if avg_daily_consumption > 0 else 999
        reorder_date = today + timedelta(days=max(days_remaining - 3, 0))  # 3-day buffer

        return ok(data={
            'variant_id': variant_id,
            'product_name': variant.product.name,
            'variant_label': f'{variant.size} {variant.flavour}'.strip(),
            'sku': variant.sku,
            'current_stock': current_stock,
            'avg_daily_consumption': round(avg_daily_consumption, 2),
            'days_remaining': days_remaining,
            'reorder_date_estimate': reorder_date.isoformat(),
            'reorder_point': variant.reorder_point,
            'history': [
                {
                    'date': d['timestamp__date'].isoformat(),
                    'out_qty': float(d['daily_out']),
                }
                for d in daily_outs
            ],
        })
